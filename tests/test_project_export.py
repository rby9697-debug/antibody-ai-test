from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.importer import _first_non_empty_row, import_project
from app.project_exporter import export_project


def _count_non_empty_data_rows(workbook_path: Path) -> dict[str, int]:
    workbook = load_workbook(workbook_path)
    counts: dict[str, int] = {}
    for ws in workbook.worksheets:
        header_row = _first_non_empty_row(ws)
        count = 0
        for row_idx in range(header_row, ws.max_row + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            if any(value not in (None, "") for value in values):
                count += 1
        counts[ws.title] = count
    return counts


def test_import_then_export_row_counts_match(tmp_path):
    db_path = tmp_path / "test.db"
    engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

    source_template = Path("docs/templates/SG866-template.xlsx")

    with TestingSessionLocal() as db:
        import_project(db, project_id=1, template_path=str(source_template))
        export_path = export_project(db, project_id=1)

    original_counts = _count_non_empty_data_rows(source_template)
    exported_counts = _count_non_empty_data_rows(export_path)

    assert original_counts == exported_counts
