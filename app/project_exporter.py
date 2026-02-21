import json
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session, selectinload

from app.models import Project, ProjectRow, ProjectSheet

BASE_TEMPLATE = Path("docs/templates/SG866-template.xlsx")


def export_project(db: Session, project_id: int) -> Path:
    project = (
        db.query(Project)
        .options(selectinload(Project.sheets).selectinload(ProjectSheet.rows))
        .filter(Project.id == project_id)
        .first()
    )
    if project is None:
        raise ValueError(f"Project {project_id} not found")

    workbook = load_workbook(BASE_TEMPLATE)

    for ws in workbook.worksheets:
        if ws.title not in {sheet.name for sheet in project.sheets}:
            ws.delete_rows(1, ws.max_row)

    sheets_by_name = {sheet.name: sheet for sheet in project.sheets}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet = sheets_by_name.get(sheet_name)
        if sheet is None:
            continue

        rows = (
            db.query(ProjectRow)
            .filter(ProjectRow.sheet_id == sheet.id)
            .order_by(ProjectRow.display_order.asc())
            .all()
        )

        ws.delete_rows(sheet.header_row, ws.max_row - sheet.header_row + 1)

        write_row = sheet.header_row
        for stored_row in rows:
            values = json.loads(stored_row.values)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=write_row, column=col_idx, value=value)
            write_row += 1

    output_path = Path(f"storage/export_{project_id}.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
