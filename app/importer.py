import json
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Project, ProjectRow, ProjectSheet


def _first_non_empty_row(worksheet) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        values = [worksheet.cell(row=row_idx, column=col).value for col in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in values):
            return row_idx
    return 1


def import_project(db: Session, project_id: int, template_path: str) -> Project:
    """Import workbook data into the database with full transaction safety."""
    try:
        workbook = load_workbook(template_path)

        project = db.get(Project, project_id)
        if project is None:
            project = Project(id=project_id, template_name=Path(template_path).name)
            db.add(project)
        else:
            project.template_name = Path(template_path).name
            db.query(ProjectRow).filter(ProjectRow.project_id == project_id).delete()
            db.query(ProjectSheet).filter(ProjectSheet.project_id == project_id).delete()

        db.flush()

        sheet_order = 1
        row_order = 1
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            header_row = _first_non_empty_row(ws)
            project_sheet = ProjectSheet(
                project_id=project_id,
                display_order=sheet_order,
                name=sheet_name,
                header_row=header_row,
            )
            db.add(project_sheet)
            db.flush()

            offset = 0
            for row_idx in range(header_row, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
                if not any(value not in (None, "") for value in row_values):
                    continue
                db.add(
                    ProjectRow(
                        project_id=project_id,
                        sheet_id=project_sheet.id,
                        display_order=row_order,
                        row_offset=offset,
                        values=json.dumps(row_values, ensure_ascii=False),
                    )
                )
                row_order += 1
                offset += 1

            sheet_order += 1

        db.commit()
        db.refresh(project)
        return project
    except Exception:
        db.rollback()
        raise
